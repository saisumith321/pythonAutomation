import openpyxl
from openpyxl.styles  import PatternFill

def Row(file,sheet):
    worbook = openpyxl.load_workbook(file)
    work = worbook[sheet]
    return work.max_row

def Col(file,sheet):
    worbook = openpyxl.load_workbook(file)
    work = worbook[sheet]
    return work.max_column

def read_data(file,sheet,row,col):
    worbook = openpyxl.load_workbook(file)
    Name = worbook[sheet]
    return Name.cell(row,col).value


def write_data(file, sheet, row, col, data) :
    worbook = openpyxl.load_workbook(file)
    Name = worbook[sheet]
    Name.cell(row, col).value = data
    worbook.save(file)


def fill_green(file,sheet,row,col):
    worbook = openpyxl.load_workbook(file)
    Name = worbook[sheet]
    fill = PatternFill(start_color='60b212',end_color='60b212',fill_type='solid')
    Name.cell(row,col).fill= fill
    worbook.save(file)

def fill_red(file,sheet,row,col):
    worbook = openpyxl.load_workbook(file)
    Name = worbook[sheet]
    fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
    Name.cell(row, col).fill = fill
    worbook.save(file)


