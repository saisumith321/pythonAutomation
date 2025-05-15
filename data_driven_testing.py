import time
from selenium import webdriver
from selenium.webdriver.chrome.webdriver import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

s = Service(r"C:\Users\chromedriver-win64\chromedriver.exe")
driver = webdriver.Chrome(service=s)

file = "C:\\Users\\harsh\\Documents\\sumith.xlsx"

a = openpyxl.load_workbook(file)
b =   a['Sheet2']
# adding the data to the excel
b.cell(1,1).value = "sumith"
b.cell(1,2).value="ejbej"
b.cell(1,3).value = 32

b.cell(2,1).value = "sumith"
b.cell(2,2).value="ejbej"
b.cell(2,3).value = 32

a.save(file)
#
# rows =  b.max_row
# columns = b.max_column
# print(rows)
# print(columns)
#reading the data from the excelsheet
# for i in range(2,rows + 1):
#     for j in range(1,columns + 1):
#         data = b.cell(i,j).value
#         if data == "Bolt" or data == "roy":
#             print(b.cell(i,2).value)
#       print()











